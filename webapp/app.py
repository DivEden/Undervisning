import os
import sys
import tempfile
import zipfile
import uuid
import calendar
from collections import defaultdict
from datetime import date

from flask import (
    Flask,
    render_template,
    request,
    send_file,
    jsonify,
    session,
)

# Gør at logic.py kan importeres uanset hvorfra serveren startes
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from tools.excel_import.logic import extract_file, list_importable_sheets, write_to_excel
from tools.template_maker.logic import create_year_template, suggest_next_year
import openpyxl

app = Flask(__name__)
# Stabil secret_key: brug miljøvariabel i produktion (Render),
# fald tilbage til en tilfældig nøgle ved lokal udvikling.
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24).hex())

# Tillad store Excel-uploads (op til 50 MB)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

# Midlertidig mappe til uploaded filer i sessionens levetid
UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "museer_webapp")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXCEL = {".xlsx"}
ALLOWED_DATA  = {".docx", ".pdf", ".zip"}


def _session_dir() -> str:
    """Returnerer (og opretter) en unik mappe per browser-session."""
    sid = session.setdefault("id", uuid.uuid4().hex)
    d = os.path.join(UPLOAD_DIR, sid)
    os.makedirs(d, exist_ok=True)
    return d


def _safe_filename(name: str) -> str:
    """Fjerner farlige tegn fra filnavne."""
    return "".join(c for c in os.path.basename(name) if c not in r'\/:*?"<>|')


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload-excel", methods=["POST"])
def upload_excel():
    """Modtager Excel-filen og returnerer liste af importerbare ark."""
    f = request.files.get("excel")
    if not f:
        return jsonify({"error": "Ingen fil modtaget"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXCEL:
        return jsonify({"error": "Kun .xlsx-filer er tilladt"}), 400

    sdir = _session_dir()
    excel_path = os.path.join(sdir, "data.xlsx")
    f.save(excel_path)

    try:
        sheets = list_importable_sheets(excel_path)
    except Exception as e:
        return jsonify({"error": f"Kunne ikke læse Excel-filen: {e}"}), 400

    if not sheets:
        return jsonify({"error": "Ingen importerbare årsark fundet i filen"}), 400

    return jsonify({"sheets": sheets, "filename": _safe_filename(f.filename)})


@app.route("/import", methods=["POST"])
def do_import():
    """
    Modtager Word/PDF-filer (eller en zip), behandler dem og
    returnerer den opdaterede Excel-fil som download.
    """
    sheet_name = request.form.get("sheet", "").strip()
    if not sheet_name:
        return jsonify({"error": "Intet ark valgt"}), 400

    sdir = _session_dir()
    excel_path = os.path.join(sdir, "data.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": "Excel-filen er ikke uploaded endnu"}), 400

    # Gem uploadede datafiler i en undermappe
    data_dir = os.path.join(sdir, "input")
    os.makedirs(data_dir, exist_ok=True)

    uploaded_files = request.files.getlist("datafiles")
    if not uploaded_files or all(f.filename == "" for f in uploaded_files):
        return jsonify({"error": "Ingen Word/PDF-filer uploaded"}), 400

    saved_paths: list[str] = []
    for f in uploaded_files:
        if not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_DATA:
            continue
        dest = os.path.join(data_dir, _safe_filename(f.filename))
        f.save(dest)

        if ext == ".zip":
            # Udpak zip og tilføj indeholdte .docx/.pdf filer
            with zipfile.ZipFile(dest) as zf:
                for member in zf.infolist():
                    mext = os.path.splitext(member.filename)[1].lower()
                    if mext in {".docx", ".pdf"}:
                        mname = _safe_filename(member.filename)
                        out = os.path.join(data_dir, mname)
                        with zf.open(member) as src, open(out, "wb") as dst:
                            dst.write(src.read())
                        saved_paths.append(out)
        else:
            saved_paths.append(dest)

    if not saved_paths:
        return jsonify({"error": "Ingen gyldige .docx eller .pdf-filer fundet"}), 400

    # Udtræk data fra alle filer
    all_rows: list[dict] = []
    file_errors: list[str] = []
    for path in sorted(saved_paths):
        rows, err = extract_file(path)
        if err:
            file_errors.append(f"{os.path.basename(path)}: {err}")
        else:
            all_rows.extend(rows)

    if not all_rows:
        msg = "Ingen data fundet."
        if file_errors:
            msg += " Fejl: " + "; ".join(file_errors)
        return jsonify({"error": msg}), 400

    # Skriv til Excel
    try:
        count, warnings = write_to_excel(all_rows, excel_path, sheet_name)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    result_info = f"{count} rækker skrevet til ark '{sheet_name}'"
    if file_errors:
        result_info += f" ({len(file_errors)} fil(er) sprunget over)"

    # Send den opdaterede fil tilbage
    original_name = request.form.get("original_filename", "data.xlsx")
    download_name = _safe_filename(original_name)

    return send_file(
        excel_path,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/ny-ark", methods=["POST"])
def ny_ark():
    """Opretter et nyt årsark (fx 2027) i Excel-filen."""
    sdir = _session_dir()
    excel_path = os.path.join(sdir, "data.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": "Excel-filen er ikke uploaded endnu"}), 400

    try:
        year = int(request.form.get("year", 0))
    except ValueError:
        return jsonify({"error": "Ugyldigt årstal"}), 400

    try:
        year_sheet, data_sheet = create_year_template(excel_path, year)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Kunne ikke oprette ark: {e}"}), 500

    original_name = request.form.get("original_filename", "data.xlsx")
    return send_file(
        excel_path,
        as_attachment=True,
        download_name=_safe_filename(original_name),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/suggest-year")
def suggest_year():
    """Returnerer foreslået næste årstal baseret på Excel-filen."""
    sdir = _session_dir()
    excel_path = os.path.join(sdir, "data.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"year": date.today().year + 1})
    try:
        year = suggest_next_year(excel_path)
        return jsonify({"year": year})
    except Exception:
        return jsonify({"year": date.today().year + 1})


# ─── Grafer ──────────────────────────────────────────────────────────────────

@app.route("/grafer")
def grafer():
    return render_template("grafer.html")


@app.route("/api/grafer-data")
def grafer_data():
    """Returnerer aggregeret data fra alle årsark som JSON til Chart.js."""
    sdir = _session_dir()
    excel_path = os.path.join(sdir, "data.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": "Ingen Excel-fil uploaded"}), 400

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    MAANEDER = ["Jan", "Feb", "Mar", "Apr", "Maj", "Jun",
                "Jul", "Aug", "Sep", "Okt", "Nov", "Dec"]

    years_data = {}   # år -> { besøg, elever, lærere, måneder, skoletype, forløb, ulf, aarhus }

    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        besog_per_maaned   = [0] * 12
        elever_per_maaned  = [0] * 12
        laerere_per_maaned = [0] * 12
        skoletype_count    = defaultdict(int)
        forlob_count       = defaultdict(int)
        ulf_count          = 0
        aarhus_count       = 0
        total_besog        = 0
        total_elever       = 0
        total_laerere      = 0

        for row in rows:
            if not row or row[0] is None:
                continue
            dato       = row[0]
            skoletype  = str(row[2]).strip() if row[2] else "Ukendt"
            elever     = int(row[4]) if isinstance(row[4], (int, float)) else 0
            laerere    = int(row[5]) if isinstance(row[5], (int, float)) else 0
            forlob     = str(row[6]).strip() if row[6] else "Ukendt"
            ulf        = row[11]
            aarhus     = row[12]

            if hasattr(dato, "month"):
                m = dato.month - 1
                besog_per_maaned[m]   += 1
                elever_per_maaned[m]  += elever
                laerere_per_maaned[m] += laerere

            skoletype_count[skoletype] += 1
            if forlob and forlob != "None":
                forlob_count[forlob] += 1
            if ulf and str(ulf).strip().lower() in ("x", "1", "true", "ja"):
                ulf_count += 1
            if aarhus and str(aarhus).strip().lower() in ("x", "1", "true", "ja"):
                aarhus_count += 1
            total_besog   += 1
            total_elever  += elever
            total_laerere += laerere

        # Top 8 undervisningsforløb
        top_forlob = sorted(forlob_count.items(), key=lambda x: -x[1])[:8]

        years_data[sheet_name] = {
            "besog_per_maaned":   besog_per_maaned,
            "elever_per_maaned":  elever_per_maaned,
            "laerere_per_maaned": laerere_per_maaned,
            "skoletype":          dict(skoletype_count),
            "top_forlob":         top_forlob,
            "ulf":                ulf_count,
            "ikke_ulf":           total_besog - ulf_count,
            "aarhus":             aarhus_count,
            "ikke_aarhus":        total_besog - aarhus_count,
            "total_besog":        total_besog,
            "total_elever":       total_elever,
            "total_laerere":      total_laerere,
        }

    wb.close()

    sorted_years = sorted(years_data.keys())
    return jsonify({
        "maaneder": MAANEDER,
        "years":    sorted_years,
        "data":     years_data,
    })

