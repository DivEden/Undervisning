from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from tools.excel_import.logic import (
    EXPECTED_DATA_HEADERS,
    _repair_missing_pivot_cache_records,
    create_excel_backup,
    is_data_sheet,
)


def _table_names(workbook) -> set[str]:
    names: set[str] = set()
    for ws in workbook.worksheets:
        for tbl in ws.tables.values():
            names.add(tbl.name)
    return names


def _next_table_name(workbook) -> str:
    used = _table_names(workbook)
    i = 1
    while True:
        name = f"Tabel{i}"
        if name not in used:
            return name
        i += 1


def _find_latest_data_sheet(workbook):
    candidates = []
    for name in workbook.sheetnames:
        if name.isdigit() and is_data_sheet(workbook[name]):
            candidates.append((int(name), name))
    if not candidates:
        return None
    candidates.sort()
    return candidates[-1][1]


def suggest_next_year(excel_path: str) -> int:
    _repair_missing_pivot_cache_records(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        years = [int(name) for name in wb.sheetnames if name.isdigit()]
        if not years:
            return 2027
        return max(years) + 1
    finally:
        wb.close()


def create_year_template(excel_path: str, year: int) -> tuple[str, str]:
    if year < 2000 or year > 2100:
        raise ValueError("År skal være mellem 2000 og 2100.")

    year_sheet_name = str(year)
    data_sheet_name = f"Data {year}"

    _repair_missing_pivot_cache_records(excel_path, create_backup_before_change=True)
    wb = openpyxl.load_workbook(excel_path)

    if year_sheet_name in wb.sheetnames:
        raise ValueError(f"Arket '{year_sheet_name}' findes allerede.")
    if data_sheet_name in wb.sheetnames:
        raise ValueError(f"Arket '{data_sheet_name}' findes allerede.")

    source_name = _find_latest_data_sheet(wb)
    source_ws = wb[source_name] if source_name else None

    create_excel_backup(excel_path, reason="template_maker")

    year_ws = wb.create_sheet(year_sheet_name)

    # Header-tekster
    for idx, header in enumerate(EXPECTED_DATA_HEADERS, start=1):
        year_ws.cell(1, idx, header)

    # Kopiér visuel stil fra seneste data-ark hvis muligt
    if source_ws is not None:
        for col in range(1, len(EXPECTED_DATA_HEADERS) + 1):
            src_cell = source_ws.cell(1, col)
            dst_cell = year_ws.cell(1, col)
            dst_cell._style = copy(src_cell._style)
            dst_cell.number_format = src_cell.number_format

            col_letter = openpyxl.utils.get_column_letter(col)
            src_dim = source_ws.column_dimensions[col_letter]
            dst_dim = year_ws.column_dimensions[col_letter]
            if src_dim.width is not None:
                dst_dim.width = src_dim.width

        if source_ws.row_dimensions[1].height is not None:
            year_ws.row_dimensions[1].height = source_ws.row_dimensions[1].height

        year_ws.freeze_panes = source_ws.freeze_panes or "A2"
    else:
        # Fallback hvis der ikke findes data-ark
        for col in range(1, len(EXPECTED_DATA_HEADERS) + 1):
            c = year_ws.cell(1, col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0F766E")
            c.alignment = Alignment(horizontal="center", vertical="center")
            year_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        year_ws.freeze_panes = "A2"

    # Opret en tom datarække så tabel-stilen (grå/hvid striber) er aktiv fra start
    for col in range(1, len(EXPECTED_DATA_HEADERS) + 1):
        year_ws.cell(2, col, None)

    style_name = "TableStyleLight1"
    if source_ws is not None and source_ws.tables:
        first_table = next(iter(source_ws.tables.values()))
        if first_table.tableStyleInfo and first_table.tableStyleInfo.name:
            style_name = first_table.tableStyleInfo.name

    tbl = Table(displayName=_next_table_name(wb), ref=f"A1:M2")
    tbl.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    year_ws.add_table(tbl)

    # Opret tilhørende data-ark til statistik
    data_ws = wb.create_sheet(data_sheet_name)
    data_ws["A1"] = data_sheet_name
    data_ws["A1"].font = Font(size=16, bold=True)

    data_ws["A3"] = "Dette ark bruges til statistik for året."
    data_ws["A4"] = "Brug værktøjet 'Statistik Maker' for at generere tabeller her."
    data_ws["A3"].font = Font(size=11)
    data_ws["A4"].font = Font(size=11)
    data_ws.column_dimensions["A"].width = 64

    wb.save(excel_path)
    _repair_missing_pivot_cache_records(excel_path, create_backup_before_change=False)

    return year_sheet_name, data_sheet_name
