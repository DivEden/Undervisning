import math
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tools.excel_import.logic import (
    EXPECTED_DATA_HEADERS,
    _repair_missing_pivot_cache_records,
    create_excel_backup,
    is_data_sheet,
)

GROUP_OPTIONS = {
    "year": "År",
    "month": "Måned",
    "skoletype": "Skoletype",
    "forloeb": "Undervisningsforløb",
    "skole": "Skole",
    "peh_flag": "PEH (x / tom)",
    "special_flag": "Specialklasse (x / tom)",
}

METRIC_OPTIONS = {
    "count_rows": "Antal besøg",
    "sum_elever": "Sum elever",
    "sum_laerere": "Sum lærere",
    "count_peh": "Antal PEH",
    "count_special": "Antal specialklasse",
    "unique_schools": "Antal unikke skoler",
    "matrix_monthly_year_visits_students": "Månedlig sammenligning (Besøg + Elever pr. år)",
}

TEMPLATE_OPTIONS = {
    "template_monthly_comparison": "Template: Månedlig sammenligning (Besøg + Elever)",
    "template_skoletype_peh_uf": "Template: Skoletype (PEH + UF + Elever + Lærere)",
    "template_forloeb_ulf": "Template: Undervisningsforløb (Antal + ULF + Elever)",
    "template_download_materiale": "Template: Downloadet materiale",
    "template_klassetrin": "Template: Klassetrin",
    "template_peh_values": "Template: PEH-værdier (x/(tom))",
    "template_full_overview": "Template: Fuld oversigt (alle tabeller)",
    "template_extra_year_overview": "Template EKSTRA: Årsoversigt (KPI)",
    "template_extra_year_yoy": "Template EKSTRA: År-over-år udvikling",
    "template_extra_skoletype_by_year": "Template EKSTRA: Skoletype pr. år (matrix)",
    "template_extra_forloeb_by_year": "Template EKSTRA: Undervisningsforløb pr. år (matrix)",
    "template_extra_aarhus_split": "Template EKSTRA: Aarhus vs. øvrige pr. år",
    "template_extra_top_schools": "Template EKSTRA: Top skoler",
    "template_extra_activity_mix": "Template EKSTRA: Aktivitetsmix pr. år",
    "template_extra_efficiency": "Template EKSTRA: Effektivitet (elever/lærere)",
    "template_extra_monthly_yoy": "Template EKSTRA: Månedlig år-over-år vækst",
    "template_extra_pack": "Template EKSTRA: Stor pakke (mange analyser)",
}

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="F3F4F6")
TOTAL_FONT = Font(bold=True)
THIN_SIDE = Side(style="thin", color="D1D5DB")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _to_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_x(value) -> bool:
    return str(value or "").strip().lower() == "x"


def _safe_label(value, empty_label="(tom)"):
    txt = "" if value is None else str(value).strip()
    return txt if txt else empty_label


def _month_label(dt: datetime | None) -> str:
    if not isinstance(dt, datetime):
        return "Uden dato"
    labels = ["jan", "feb", "mar", "apr", "maj", "jun", "jul", "aug", "sep", "okt", "nov", "dec"]
    return labels[dt.month - 1]


def _month_full_label(month_key: str) -> str:
    mapping = {
        "jan": "Januar",
        "feb": "Februar",
        "mar": "Marts",
        "apr": "April",
        "maj": "Maj",
        "jun": "Juni",
        "jul": "Juli",
        "aug": "August",
        "sep": "September",
        "okt": "Oktober",
        "nov": "November",
        "dec": "December",
    }
    return mapping.get(month_key, month_key)


def list_data_sheets(excel_path: str) -> list[str]:
    _repair_missing_pivot_cache_records(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        return [name for name in wb.sheetnames if wb[name].sheet_state == "visible" and is_data_sheet(wb[name])]
    finally:
        wb.close()


def list_target_sheets(excel_path: str) -> list[str]:
    _repair_missing_pivot_cache_records(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        return [name for name in wb.sheetnames if wb[name].sheet_state == "visible"]
    finally:
        wb.close()


def _rows_from_sheet(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        rows.append({
            "Dato": row[0],
            "Skole": row[1],
            "Skoletype": row[2],
            "Klassetrin": row[3],
            "Elever": row[4],
            "Lærere": row[5],
            "Undervisningsforløb": row[6],
            "Rundvisning": row[7],
            "PEH": row[8],
            "Downloadet Materiale": row[9],
            "Specialklasse": row[10],
            "ULF": row[11],
            "Aarhus Kommune": row[12] if len(row) >= 13 else None,
        })
    return rows


def _collect_rows_for_source(wb, source_choice: str):
    data_sheet_names = [name for name in wb.sheetnames if is_data_sheet(wb[name])]
    if not data_sheet_names:
        raise ValueError("Ingen data-ark fundet i Excel-filen.")

    rows = []
    if source_choice == "__ALL_DATA_SHEETS__":
        for name in data_sheet_names:
            rows.extend(_rows_from_sheet(wb[name]))
        source_text = "Alle data-ark"
        source_slug = "all"
    else:
        if source_choice not in data_sheet_names:
            raise ValueError("Valgt kildeark er ikke et data-ark.")
        rows = _rows_from_sheet(wb[source_choice])
        source_text = source_choice
        source_slug = source_choice.lower().replace(" ", "_")

    if not rows:
        raise ValueError("Ingen rækker fundet i valgt kilde.")

    return rows, source_text, source_slug


def _group_key(row: dict, group_by: str):
    if group_by == "year":
        dt = row.get("Dato")
        return str(dt.year) if isinstance(dt, datetime) else "Uden dato"
    if group_by == "month":
        return _month_label(row.get("Dato"))
    if group_by == "skoletype":
        return _safe_label(row.get("Skoletype"))
    if group_by == "forloeb":
        return _safe_label(row.get("Undervisningsforløb"))
    if group_by == "skole":
        return _safe_label(row.get("Skole"))
    if group_by == "peh_flag":
        return "x" if _is_x(row.get("PEH")) else "(tom)"
    if group_by == "special_flag":
        return "x" if _is_x(row.get("Specialklasse")) else "(tom)"
    return "Ukendt"


def _metric_value(row: dict, metric: str):
    if metric == "count_rows":
        return 1
    if metric == "sum_elever":
        return _to_float(row.get("Elever"))
    if metric == "sum_laerere":
        return _to_float(row.get("Lærere"))
    if metric == "count_peh":
        return 1 if _is_x(row.get("PEH")) else 0
    if metric == "count_special":
        return 1 if _is_x(row.get("Specialklasse")) else 0
    if metric == "unique_schools":
        return _safe_label(row.get("Skole"), empty_label="")
    return 0


def _aggregate(rows: list[dict], group_by: str, metric: str):
    if metric == "unique_schools":
        grouped = defaultdict(set)
        for row in rows:
            key = _group_key(row, group_by)
            school = _metric_value(row, metric)
            if school:
                grouped[key].add(school)
        result = [(k, len(v)) for k, v in grouped.items()]
    else:
        grouped = defaultdict(float)
        for row in rows:
            key = _group_key(row, group_by)
            grouped[key] += _metric_value(row, metric)
        result = [(k, v) for k, v in grouped.items()]

    if group_by == "month":
        order = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "maj": 5, "jun": 6, "jul": 7, "aug": 8, "sep": 9, "okt": 10, "nov": 11, "dec": 12, "Uden dato": 13}
        result.sort(key=lambda item: order.get(item[0], 99))
    else:
        result.sort(key=lambda item: str(item[0]).lower())
    return result


def _aggregate_monthly_year_matrix(rows: list[dict]):
    month_order = ["jan", "feb", "mar", "apr", "maj", "jun", "jul", "aug", "sep", "okt", "nov", "dec"]

    years = sorted(
        {
            row["Dato"].year
            for row in rows
            if isinstance(row.get("Dato"), datetime)
        }
    )
    if not years:
        raise ValueError("Ingen gyldige datoer fundet til månedlig sammenligning.")

    stats = defaultdict(lambda: {"visits": 0, "students": 0.0})
    for row in rows:
        dt = row.get("Dato")
        if not isinstance(dt, datetime):
            continue
        mk = _month_label(dt)
        y = dt.year
        stats[(mk, y)]["visits"] += 1
        stats[(mk, y)]["students"] += _to_float(row.get("Elever"))

    headers = ["Måned"]
    for y in years:
        headers.extend([f"Besøg {y}", f"Elever {y}"])

    rows_out = []
    for mk in month_order:
        has_data = any(stats[(mk, y)]["visits"] > 0 for y in years)
        if not has_data:
            continue

        row_vals = [_month_full_label(mk)]
        for y in years:
            visits = stats[(mk, y)]["visits"]
            students = round(stats[(mk, y)]["students"])
            row_vals.extend([
                visits if visits > 0 else None,
                students if students > 0 else None,
            ])
        rows_out.append(row_vals)

    total_row = ["TOTAL"]
    for y in years:
        total_visits = sum(stats[(mk, y)]["visits"] for mk in month_order)
        total_students = round(sum(stats[(mk, y)]["students"] for mk in month_order))
        total_row.extend([total_visits if total_visits > 0 else None, total_students if total_students > 0 else None])
    rows_out.append(total_row)

    return headers, rows_out


def _template_stats(rows: list[dict]):
    by_type = defaultdict(lambda: {"peh": 0, "uf": 0, "elever": 0.0, "laerere": 0.0, "ulf": 0, "klassetrin": 0})
    by_forloeb = defaultdict(lambda: {"antal": 0, "ulf": 0, "elever": 0.0})
    by_mat = defaultdict(int)
    by_klassetrin = defaultdict(int)
    by_peh = defaultdict(int)

    for row in rows:
        stype = _safe_label(row.get("Skoletype"))
        forloeb = _safe_label(row.get("Undervisningsforløb"), empty_label="")
        klassetrin = _safe_label(row.get("Klassetrin"), empty_label="")
        mat = _safe_label(row.get("Downloadet Materiale"), empty_label="")

        elever = _to_float(row.get("Elever"))
        laerere = _to_float(row.get("Lærere"))
        peh_is_x = _is_x(row.get("PEH"))
        ulf_is_x = _is_x(row.get("ULF"))

        by_type[stype]["elever"] += elever
        by_type[stype]["laerere"] += laerere
        if peh_is_x:
            by_type[stype]["peh"] += 1
        if ulf_is_x:
            by_type[stype]["ulf"] += 1
        if forloeb:
            by_type[stype]["uf"] += 1
        if klassetrin:
            by_type[stype]["klassetrin"] += 1

        if forloeb:
            by_forloeb[forloeb]["antal"] += 1
            by_forloeb[forloeb]["elever"] += elever
            if ulf_is_x:
                by_forloeb[forloeb]["ulf"] += 1

        if mat and mat.lower() != "nej":
            by_mat[mat] += 1

        if klassetrin:
            by_klassetrin[klassetrin] += 1

        by_peh["x" if peh_is_x else "(tom)"] += 1

    return by_type, by_forloeb, by_mat, by_klassetrin, by_peh


def _extract_years(rows: list[dict]) -> list[int]:
    return sorted({r["Dato"].year for r in rows if isinstance(r.get("Dato"), datetime)})


def _percent(part: float, whole: float) -> float | None:
    if whole <= 0:
        return None
    return round((part / whole) * 100, 1)


def _with_total(headers: list[str], rows: list[list], total_label: str = "Hovedtotal"):
    if not rows:
        return headers, rows

    total = [total_label]
    for col in range(1, len(headers)):
        header_name = str(headers[col]).lower()
        if "%" in header_name:
            total.append(None)
            continue

        col_sum = 0
        has_numeric = False
        has_fraction = False
        for row in rows:
            value = row[col]
            if isinstance(value, (int, float)):
                col_sum += value
                has_numeric = True
                if isinstance(value, float) and not math.isclose(value, int(value), rel_tol=0, abs_tol=1e-9):
                    has_fraction = True

        if not has_numeric:
            total.append(None)
        elif has_fraction:
            total.append(round(col_sum, 2))
        else:
            total.append(int(round(col_sum)))

    return headers, rows + [total]


def _build_template_blocks(rows: list[dict], template_id: str):
    by_type, by_forloeb, by_mat, by_klassetrin, by_peh = _template_stats(rows)

    monthly_headers, monthly_rows = _aggregate_monthly_year_matrix(rows)

    type_headers = ["Rækkemærkater", "Antal af PEH", "Antal af Undervisningsforløb", "Sum af Elever", "Sum af Lærere"]
    type_rows = []
    for key, v in sorted(by_type.items(), key=lambda item: (-item[1]["peh"], item[0].lower())):
        type_rows.append([
            key,
            v["peh"] if v["peh"] > 0 else None,
            v["uf"] if v["uf"] > 0 else None,
            int(round(v["elever"])) if v["elever"] > 0 else None,
            round(v["laerere"], 1) if v["laerere"] > 0 else None,
        ])
    type_headers, type_rows = _with_total(type_headers, type_rows)

    forloeb_headers = ["Rækkemærkater", "Antal af Undervisningsforløb", "Antal af ULF", "Sum af Elever"]
    forloeb_rows = []
    for key, v in sorted(by_forloeb.items(), key=lambda item: (-item[1]["antal"], item[0].lower())):
        forloeb_rows.append([
            key,
            v["antal"] if v["antal"] > 0 else None,
            v["ulf"] if v["ulf"] > 0 else None,
            int(round(v["elever"])) if v["elever"] > 0 else None,
        ])
    forloeb_headers, forloeb_rows = _with_total(forloeb_headers, forloeb_rows)

    materiale_headers = ["Rækkemærkater", "Antal af Downloadet Materiale"]
    materiale_rows = [[k, v] for k, v in sorted(by_mat.items(), key=lambda item: (-item[1], item[0].lower()))]
    materiale_headers, materiale_rows = _with_total(materiale_headers, materiale_rows)

    klassetrin_headers = ["Rækkemærkater", "Antal af Klassetrin"]
    klassetrin_rows = [[k, v] for k, v in sorted(by_klassetrin.items(), key=lambda item: (-item[1], item[0].lower()))]
    klassetrin_headers, klassetrin_rows = _with_total(klassetrin_headers, klassetrin_rows)

    peh_headers = ["Rækkemærkater", "Antal af PEH"]
    peh_rows = [[k, v] for k, v in sorted(by_peh.items(), key=lambda item: (-item[1], item[0].lower()))]
    peh_headers, peh_rows = _with_total(peh_headers, peh_rows)

    years = _extract_years(rows)
    if not years:
        raise ValueError("Ingen gyldige år fundet i data.")

    yearly = defaultdict(
        lambda: {
            "visits": 0,
            "students": 0.0,
            "teachers": 0.0,
            "peh": 0,
            "ulf": 0,
            "forloeb": 0,
            "download": 0,
            "special": 0,
            "schools": set(),
            "aarhus": 0,
            "other": 0,
        }
    )
    schools = defaultdict(lambda: {"visits": 0, "students": 0.0, "years": defaultdict(int)})
    by_type_year = defaultdict(lambda: defaultdict(int))
    by_forloeb_year = defaultdict(lambda: defaultdict(int))

    for row in rows:
        dt = row.get("Dato")
        if not isinstance(dt, datetime):
            continue
        y = dt.year
        stype = _safe_label(row.get("Skoletype"))
        school = _safe_label(row.get("Skole"))
        forloeb = _safe_label(row.get("Undervisningsforløb"), empty_label="")
        mat = _safe_label(row.get("Downloadet Materiale"), empty_label="")

        yearly[y]["visits"] += 1
        yearly[y]["students"] += _to_float(row.get("Elever"))
        yearly[y]["teachers"] += _to_float(row.get("Lærere"))
        yearly[y]["schools"].add(school)

        if _is_x(row.get("PEH")):
            yearly[y]["peh"] += 1
        if _is_x(row.get("ULF")):
            yearly[y]["ulf"] += 1
        if _is_x(row.get("Specialklasse")):
            yearly[y]["special"] += 1
        if forloeb:
            yearly[y]["forloeb"] += 1
            by_forloeb_year[forloeb][y] += 1
        if mat and mat.lower() != "nej":
            yearly[y]["download"] += 1

        if _is_x(row.get("Aarhus Kommune")):
            yearly[y]["aarhus"] += 1
        else:
            yearly[y]["other"] += 1

        by_type_year[stype][y] += 1

        schools[school]["visits"] += 1
        schools[school]["students"] += _to_float(row.get("Elever"))
        schools[school]["years"][y] += 1

    # EKSTRA: Årsoversigt KPI
    year_overview_headers = [
        "År",
        "Besøg",
        "Elever",
        "Lærere",
        "PEH",
        "ULF",
        "Unikke skoler",
        "PEH-andel %",
    ]
    year_overview_rows = []
    for y in years:
        d = yearly[y]
        year_overview_rows.append(
            [
                y,
                d["visits"],
                int(round(d["students"])),
                round(d["teachers"], 1),
                d["peh"],
                d["ulf"],
                len(d["schools"]),
                _percent(d["peh"], d["visits"]),
            ]
        )
    year_overview_headers, year_overview_rows = _with_total(year_overview_headers, year_overview_rows)

    # EKSTRA: År-over-år
    yoy_headers = ["År", "Besøg", "Delta besøg", "Elever", "Delta elever", "PEH-andel %", "ULF-andel %"]
    yoy_rows = []
    prev_visits = None
    prev_students = None
    for y in years:
        d = yearly[y]
        visits = d["visits"]
        students = int(round(d["students"]))
        yoy_rows.append(
            [
                y,
                visits,
                (visits - prev_visits) if prev_visits is not None else None,
                students,
                (students - prev_students) if prev_students is not None else None,
                _percent(d["peh"], visits),
                _percent(d["ulf"], visits),
            ]
        )
        prev_visits = visits
        prev_students = students
    yoy_headers, yoy_rows = _with_total(yoy_headers, yoy_rows)

    # EKSTRA: Skoletype x År
    type_year_headers = ["Skoletype"]
    for y in years:
        type_year_headers.extend([f"Besøg {y}", f"Andel {y}%"])
    type_year_rows = []
    for st in sorted(by_type_year.keys(), key=lambda k: (-sum(by_type_year[k].values()), k.lower())):
        row_vals = [st]
        for y in years:
            visits = by_type_year[st][y]
            row_vals.extend([visits if visits > 0 else None, _percent(visits, yearly[y]["visits"])])
        type_year_rows.append(row_vals)
    type_year_headers, type_year_rows = _with_total(type_year_headers, type_year_rows)

    # EKSTRA: Undervisningsforløb x År (top 20)
    top_forloeb = sorted(by_forloeb_year.keys(), key=lambda k: -sum(by_forloeb_year[k].values()))[:20]
    forloeb_year_headers = ["Undervisningsforløb"] + [str(y) for y in years] + ["I alt"]
    forloeb_year_rows = []
    for fl in top_forloeb:
        vals = [by_forloeb_year[fl][y] if by_forloeb_year[fl][y] > 0 else None for y in years]
        total = sum(by_forloeb_year[fl][y] for y in years)
        forloeb_year_rows.append([fl] + vals + [total])
    forloeb_year_headers, forloeb_year_rows = _with_total(forloeb_year_headers, forloeb_year_rows)

    # EKSTRA: Aarhus split
    aarhus_headers = ["År", "Aarhus besøg", "Øvrige besøg", "Aarhus-andel %"]
    aarhus_rows = []
    for y in years:
        a = yearly[y]["aarhus"]
        o = yearly[y]["other"]
        aarhus_rows.append([y, a if a > 0 else None, o if o > 0 else None, _percent(a, a + o)])
    aarhus_headers, aarhus_rows = _with_total(aarhus_headers, aarhus_rows)

    # EKSTRA: Top skoler (top 25)
    school_headers = ["Skole", "Besøg", "Elever", "Snit elever/besøg", "Top-år"]
    school_rows = []
    top_schools = sorted(schools.keys(), key=lambda k: -schools[k]["visits"])[:25]
    for s in top_schools:
        d = schools[s]
        top_year = max(d["years"], key=d["years"].get) if d["years"] else None
        avg = round(d["students"] / d["visits"], 1) if d["visits"] else None
        school_rows.append([s, d["visits"], int(round(d["students"])), avg, top_year])
    school_headers, school_rows = _with_total(school_headers, school_rows)

    # EKSTRA: Aktivitetsmix
    mix_headers = ["År", "Med forløb", "Med PEH", "Med ULF", "Med download", "Med specialklasse"]
    mix_rows = []
    for y in years:
        d = yearly[y]
        mix_rows.append(
            [
                y,
                d["forloeb"] if d["forloeb"] > 0 else None,
                d["peh"] if d["peh"] > 0 else None,
                d["ulf"] if d["ulf"] > 0 else None,
                d["download"] if d["download"] > 0 else None,
                d["special"] if d["special"] > 0 else None,
            ]
        )
    mix_headers, mix_rows = _with_total(mix_headers, mix_rows)

    # EKSTRA: Effektivitet
    eff_headers = ["År", "Elever", "Lærere", "Elever pr. lærer", "Elever pr. besøg"]
    eff_rows = []
    for y in years:
        d = yearly[y]
        elev = int(round(d["students"]))
        laer = d["teachers"]
        eff_rows.append(
            [
                y,
                elev,
                round(laer, 1),
                round(elev / laer, 2) if laer > 0 else None,
                round(elev / d["visits"], 2) if d["visits"] > 0 else None,
            ]
        )
    eff_headers, eff_rows = _with_total(eff_headers, eff_rows)

    # EKSTRA: Månedlig år-over-år vækst (kun de to sidste år)
    month_growth_headers = ["Måned", "Besøg sidste år", "Besøg i år", "Vækst besøg %", "Elever sidste år", "Elever i år", "Vækst elever %"]
    month_growth_rows = []
    if len(years) >= 2:
        y_prev = years[-2]
        y_curr = years[-1]
        m_headers, m_rows = _aggregate_monthly_year_matrix(rows)
        idx_prev_vis = m_headers.index(f"Besøg {y_prev}")
        idx_prev_ele = m_headers.index(f"Elever {y_prev}")
        idx_curr_vis = m_headers.index(f"Besøg {y_curr}")
        idx_curr_ele = m_headers.index(f"Elever {y_curr}")
        for r in m_rows:
            if str(r[0]).strip().upper() == "TOTAL":
                continue
            pv = r[idx_prev_vis] or 0
            pe = r[idx_prev_ele] or 0
            cv = r[idx_curr_vis] or 0
            ce = r[idx_curr_ele] or 0
            month_growth_rows.append(
                [
                    r[0],
                    pv if pv > 0 else None,
                    cv if cv > 0 else None,
                    _percent(cv - pv, pv) if pv > 0 else None,
                    pe if pe > 0 else None,
                    ce if ce > 0 else None,
                    _percent(ce - pe, pe) if pe > 0 else None,
                ]
            )
    month_growth_headers, month_growth_rows = _with_total(month_growth_headers, month_growth_rows)

    blocks = {
        "monthly": {
            "title": "Template: Månedlig sammenligning (Besøg + Elever)",
            "headers": monthly_headers,
            "rows": monthly_rows,
        },
        "skoletype": {
            "title": "Template: Skoletype (PEH + UF + Elever + Lærere)",
            "headers": type_headers,
            "rows": type_rows,
        },
        "forloeb": {
            "title": "Template: Undervisningsforløb (Antal + ULF + Elever)",
            "headers": forloeb_headers,
            "rows": forloeb_rows,
        },
        "materiale": {
            "title": "Template: Downloadet materiale",
            "headers": materiale_headers,
            "rows": materiale_rows,
        },
        "klassetrin": {
            "title": "Template: Klassetrin",
            "headers": klassetrin_headers,
            "rows": klassetrin_rows,
        },
        "peh": {
            "title": "Template: PEH-værdier (x/(tom))",
            "headers": peh_headers,
            "rows": peh_rows,
        },
        "extra_year_overview": {
            "title": "Template EKSTRA: Årsoversigt (KPI)",
            "headers": year_overview_headers,
            "rows": year_overview_rows,
        },
        "extra_yoy": {
            "title": "Template EKSTRA: År-over-år udvikling",
            "headers": yoy_headers,
            "rows": yoy_rows,
        },
        "extra_type_year": {
            "title": "Template EKSTRA: Skoletype pr. år (matrix)",
            "headers": type_year_headers,
            "rows": type_year_rows,
        },
        "extra_forloeb_year": {
            "title": "Template EKSTRA: Undervisningsforløb pr. år (matrix)",
            "headers": forloeb_year_headers,
            "rows": forloeb_year_rows,
        },
        "extra_aarhus": {
            "title": "Template EKSTRA: Aarhus vs. øvrige pr. år",
            "headers": aarhus_headers,
            "rows": aarhus_rows,
        },
        "extra_top_schools": {
            "title": "Template EKSTRA: Top skoler",
            "headers": school_headers,
            "rows": school_rows,
        },
        "extra_mix": {
            "title": "Template EKSTRA: Aktivitetsmix pr. år",
            "headers": mix_headers,
            "rows": mix_rows,
        },
        "extra_eff": {
            "title": "Template EKSTRA: Effektivitet (elever/lærere)",
            "headers": eff_headers,
            "rows": eff_rows,
        },
        "extra_month_growth": {
            "title": "Template EKSTRA: Månedlig år-over-år vækst",
            "headers": month_growth_headers,
            "rows": month_growth_rows,
        },
    }

    if template_id == "template_monthly_comparison":
        return [("monthly", blocks["monthly"])]
    if template_id == "template_skoletype_peh_uf":
        return [("skoletype", blocks["skoletype"])]
    if template_id == "template_forloeb_ulf":
        return [("forloeb", blocks["forloeb"])]
    if template_id == "template_download_materiale":
        return [("materiale", blocks["materiale"])]
    if template_id == "template_klassetrin":
        return [("klassetrin", blocks["klassetrin"])]
    if template_id == "template_peh_values":
        return [("peh", blocks["peh"])]
    if template_id == "template_full_overview":
        return [
            ("monthly", blocks["monthly"]),
            ("skoletype", blocks["skoletype"]),
            ("forloeb", blocks["forloeb"]),
            ("materiale", blocks["materiale"]),
            ("klassetrin", blocks["klassetrin"]),
            ("peh", blocks["peh"]),
        ]
    if template_id == "template_extra_year_overview":
        return [("extra_year_overview", blocks["extra_year_overview"])]
    if template_id == "template_extra_year_yoy":
        return [("extra_yoy", blocks["extra_yoy"])]
    if template_id == "template_extra_skoletype_by_year":
        return [("extra_type_year", blocks["extra_type_year"])]
    if template_id == "template_extra_forloeb_by_year":
        return [("extra_forloeb_year", blocks["extra_forloeb_year"])]
    if template_id == "template_extra_aarhus_split":
        return [("extra_aarhus", blocks["extra_aarhus"])]
    if template_id == "template_extra_top_schools":
        return [("extra_top_schools", blocks["extra_top_schools"])]
    if template_id == "template_extra_activity_mix":
        return [("extra_mix", blocks["extra_mix"])]
    if template_id == "template_extra_efficiency":
        return [("extra_eff", blocks["extra_eff"])]
    if template_id == "template_extra_monthly_yoy":
        return [("extra_month_growth", blocks["extra_month_growth"])]
    if template_id == "template_extra_pack":
        return [
            ("extra_year_overview", blocks["extra_year_overview"]),
            ("extra_yoy", blocks["extra_yoy"]),
            ("extra_aarhus", blocks["extra_aarhus"]),
            ("extra_mix", blocks["extra_mix"]),
            ("extra_eff", blocks["extra_eff"]),
            ("extra_type_year", blocks["extra_type_year"]),
            ("extra_forloeb_year", blocks["extra_forloeb_year"]),
            ("extra_top_schools", blocks["extra_top_schools"]),
            ("extra_month_growth", blocks["extra_month_growth"]),
        ]

    raise ValueError("Ukendt template.")


def _last_used_row(ws) -> int:
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 0


def _marker(block_id: str, kind: str) -> str:
    return f"__STAT_{kind}__:{block_id}"


def _remove_existing_block(ws, block_id: str):
    start_marker = _marker(block_id, "START")
    end_marker = _marker(block_id, "END")

    start_row = None
    end_row = None
    for r in range(1, ws.max_row + 1):
        value = ws.cell(r, 1).value
        if value == start_marker:
            start_row = r
        elif value == end_marker and start_row is not None:
            end_row = r
            break

    if start_row is not None and end_row is not None and end_row >= start_row:
        ws.delete_rows(start_row, end_row - start_row + 1)


def _write_block(ws, block_id: str, title: str, column_1_title: str, column_2_title: str, data_rows: list[tuple], source_text: str):
    _remove_existing_block(ws, block_id)

    start_row = _last_used_row(ws) + 2
    if start_row < 2:
        start_row = 2

    # Start marker
    ws.cell(start_row, 1, _marker(block_id, "START")).font = Font(color="FFFFFF", size=1)

    title_row = start_row + 1
    header_row = start_row + 2
    current = start_row + 3

    ws.cell(title_row, 1, title).font = Font(bold=True, size=14)
    ws.cell(title_row, 3, f"Kilde: {source_text}").font = Font(italic=True, size=10)

    h1 = ws.cell(header_row, 1, column_1_title)
    h2 = ws.cell(header_row, 2, column_2_title)
    for cell in (h1, h2):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for key, value in data_rows:
        c1 = ws.cell(current, 1, key)
        c2 = ws.cell(current, 2, value)
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        if isinstance(value, (float, int)) and not math.isclose(value, int(value), rel_tol=0, abs_tol=1e-9):
            c2.number_format = "0.00"
        current += 1

    total_row = current
    ws.cell(total_row, 1, "TOTAL")
    ws.cell(total_row, 2, f"=SUM(B{header_row + 1}:B{total_row - 1})")
    for col in (1, 2):
        cell = ws.cell(total_row, col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right" if col == 2 else "left", vertical="center")

    end_row = total_row + 1
    ws.cell(end_row, 1, _marker(block_id, "END")).font = Font(color="FFFFFF", size=1)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 36)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 16)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 40)


def _write_matrix_block(ws, block_id: str, title: str, headers: list[str], data_rows: list[list], source_text: str):
    _remove_existing_block(ws, block_id)

    start_row = _last_used_row(ws) + 2
    if start_row < 2:
        start_row = 2

    ws.cell(start_row, 1, _marker(block_id, "START")).font = Font(color="FFFFFF", size=1)

    title_row = start_row + 1
    header_row = start_row + 2
    current = start_row + 3

    ws.cell(title_row, 1, title).font = Font(bold=True, size=14)
    ws.cell(title_row, 3, f"Kilde: {source_text}").font = Font(italic=True, size=10)

    for i, label in enumerate(headers, start=1):
        h = ws.cell(header_row, i, label)
        h.fill = HEADER_FILL
        h.font = HEADER_FONT
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.border = THIN_BORDER

    for row_values in data_rows:
        for i, value in enumerate(row_values, start=1):
            c = ws.cell(current, i, value)
            c.border = THIN_BORDER
            if i == 1:
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")
        if str(row_values[0]).strip().upper() == "TOTAL":
            for i in range(1, len(headers) + 1):
                c = ws.cell(current, i)
                c.fill = TOTAL_FILL
                c.font = TOTAL_FONT
        current += 1

    end_row = current
    ws.cell(end_row, 1, _marker(block_id, "END")).font = Font(color="FFFFFF", size=1)

    for i, label in enumerate(headers, start=1):
        letter = get_column_letter(i)
        min_width = 16 if i == 1 else 13
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, max(min_width, len(str(label)) + 2))


def generate_statistics(
    excel_path: str,
    source_choice: str,
    target_sheet: str,
    group_by: str,
    metric: str,
) -> tuple[int, str]:
    if group_by not in GROUP_OPTIONS:
        raise ValueError("Ukendt gruppering.")
    if metric not in METRIC_OPTIONS:
        raise ValueError("Ukendt måling.")

    _repair_missing_pivot_cache_records(excel_path, create_backup_before_change=True)
    wb = openpyxl.load_workbook(excel_path)

    if target_sheet not in wb.sheetnames:
        raise ValueError(f"Ark '{target_sheet}' findes ikke.")

    rows, source_text, source_slug = _collect_rows_for_source(wb, source_choice)

    target_ws = wb[target_sheet]
    block_id = f"{source_slug}|{group_by}|{metric}|{target_sheet.lower().replace(' ', '_')}"

    create_excel_backup(excel_path, reason="statistics_maker")

    if metric == "matrix_monthly_year_visits_students":
        headers, matrix_rows = _aggregate_monthly_year_matrix(rows)
        title = "Statistik: Månedlig sammenligning (Besøg + Elever pr. år)"
        _write_matrix_block(
            ws=target_ws,
            block_id=block_id,
            title=title,
            headers=headers,
            data_rows=matrix_rows,
            source_text=source_text,
        )
        rows_written = len(matrix_rows)
    else:
        data_rows = _aggregate(rows, group_by, metric)
        title = f"Statistik: {METRIC_OPTIONS[metric]} pr. {GROUP_OPTIONS[group_by]}"
        _write_block(
            ws=target_ws,
            block_id=block_id,
            title=title,
            column_1_title=GROUP_OPTIONS[group_by],
            column_2_title=METRIC_OPTIONS[metric],
            data_rows=data_rows,
            source_text=source_text,
        )
        rows_written = len(data_rows)

    wb.save(excel_path)
    _repair_missing_pivot_cache_records(excel_path, create_backup_before_change=False)
    return rows_written, title


def preview_statistics(
    excel_path: str,
    source_choice: str,
    group_by: str,
    metric: str,
    max_rows: int = 14,
) -> tuple[list[str], list[list]]:
    if group_by not in GROUP_OPTIONS:
        raise ValueError("Ukendt gruppering.")
    if metric not in METRIC_OPTIONS:
        raise ValueError("Ukendt måling.")

    _repair_missing_pivot_cache_records(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        rows, _, _ = _collect_rows_for_source(wb, source_choice)

        if metric == "matrix_monthly_year_visits_students":
            headers, matrix_rows = _aggregate_monthly_year_matrix(rows)
            return headers, matrix_rows[:max_rows]

        headers = [GROUP_OPTIONS[group_by], METRIC_OPTIONS[metric]]
        data_rows = _aggregate(rows, group_by, metric)
        preview_rows = [[k, v] for k, v in data_rows[: max_rows - 1]]

        if metric == "unique_schools":
            total = sum(int(v) for _, v in data_rows)
        else:
            total = sum(float(v) for _, v in data_rows)
            if math.isclose(total, int(total), rel_tol=0, abs_tol=1e-9):
                total = int(total)

        preview_rows.append(["TOTAL", total])
        return headers, preview_rows
    finally:
        wb.close()


def generate_template(
    excel_path: str,
    source_choice: str,
    target_sheet: str,
    template_id: str,
) -> tuple[int, str]:
    if template_id not in TEMPLATE_OPTIONS:
        raise ValueError("Ukendt template.")

    _repair_missing_pivot_cache_records(excel_path, create_backup_before_change=True)
    wb = openpyxl.load_workbook(excel_path)

    if target_sheet not in wb.sheetnames:
        raise ValueError(f"Ark '{target_sheet}' findes ikke.")

    rows, source_text, source_slug = _collect_rows_for_source(wb, source_choice)
    target_ws = wb[target_sheet]
    target_slug = target_sheet.lower().replace(" ", "_")

    create_excel_backup(excel_path, reason="statistics_maker")

    blocks = _build_template_blocks(rows, template_id)
    written_rows = 0
    for block_key, block in blocks:
        block_id = f"{source_slug}|{template_id}|{block_key}|{target_slug}"
        _write_matrix_block(
            ws=target_ws,
            block_id=block_id,
            title=block["title"],
            headers=block["headers"],
            data_rows=block["rows"],
            source_text=source_text,
        )
        written_rows += len(block["rows"])

    wb.save(excel_path)
    _repair_missing_pivot_cache_records(excel_path, create_backup_before_change=False)
    return written_rows, TEMPLATE_OPTIONS[template_id]


def preview_template(
    excel_path: str,
    source_choice: str,
    template_id: str,
    max_rows_per_block: int = 8,
) -> list[dict]:
    if template_id not in TEMPLATE_OPTIONS:
        raise ValueError("Ukendt template.")

    _repair_missing_pivot_cache_records(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        rows, _, _ = _collect_rows_for_source(wb, source_choice)
        blocks = _build_template_blocks(rows, template_id)
        out = []
        for _, block in blocks:
            out.append(
                {
                    "title": block["title"],
                    "headers": block["headers"],
                    "rows": block["rows"][:max_rows_per_block],
                }
            )
        return out
    finally:
        wb.close()
